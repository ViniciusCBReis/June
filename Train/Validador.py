import pandas as pd
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.preprocessing import LabelEncoder
import joblib
import os
from openpyxl import Workbook

# Carregar os dados e pré-processamento (se necessário)
data = pd.read_csv('D:/Atividades Faculdade/APS/dados/planilhaTreino4.csv', encoding="utf8")
data['Legenda'].fillna('', inplace=True)  # Preenche os valores nulos na coluna 'Legenda' com uma string vazia
vectorizer = CountVectorizer()
X_vectorized = vectorizer.fit_transform(data['Legenda'])
label_encoder = LabelEncoder()
y_encoded = label_encoder.fit_transform(data['Relevancia'])

# Carregar o modelo treinado
modelo_carregado = joblib.load('D:/Atividades Faculdade/APS/Train/nt4.joblib')

# Função para classificar as mensagens dos usuários
def classificar_mensagem(mensagem):
    mensagem_preprocessada = vectorizer.transform([mensagem])
    intencao_predita = label_encoder.inverse_transform(modelo_carregado.predict(mensagem_preprocessada))
    return intencao_predita[0]

def criarTabela(diretorio_entrada) : 
    # Diretório com os arquivos de entrada
    # Criar uma nova pasta de trabalho
    wb = Workbook()
    # Selecionar a planilha ativa
    ws = wb.active
    # Adicionando um título para a coluna
    ws['A1'] = 'Relevancia'
    # Inicializando o contador de linha
    linha = 2
    # Iterando sobre os arquivos na pasta de entrada
    for arquivo in os.listdir(diretorio_entrada):
        if arquivo.endswith('.txt'):
            # Caminho completo do arquivo de entrada
            caminho_arquivo_entrada = os.path.join(diretorio_entrada, arquivo)
            with open(caminho_arquivo_entrada, 'r', encoding='utf-8') as file:
                # Lendo todo o conteúdo do arquivo em uma única linha
                texto = file.read()
                mensagem_usuario = texto
                intencao_predita = classificar_mensagem(mensagem_usuario)
                # Adicionando o conteúdo do arquivo como uma linha na planilha
                ws.cell(row=linha, column=1, value=intencao_predita.strip())
                # Incrementando o contador de linha para a próxima linha
                linha += 1

    wb.save('PlanilhaTreinoInicialTeste.xlsx')
    print('Concluído')

def avaliarmulti(diretorio_entrada):
    positivo = 0
    irrelevante = 0
    negativo = 0
    for arquivo in os.listdir(diretorio_entrada):
        if arquivo.endswith('.txt'):
            caminho_arquivo_entrada = os.path.join(diretorio_entrada, arquivo)
            with open(caminho_arquivo_entrada, 'r', encoding='utf-8') as file:
                # Lendo todo o conteúdo do arquivo em uma única linha
                texto = file.read()
                mensagem_usuario = texto
                intencao_predita = classificar_mensagem(mensagem_usuario)
                if intencao_predita == 'positivo':
                    positivo += 1
                elif intencao_predita == 'irrelevante':
                    irrelevante += 1
                else:
                    negativo += 1
    return positivo,irrelevante,negativo  
