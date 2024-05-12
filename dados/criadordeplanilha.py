import os
from openpyxl import Workbook

# Diretório com os arquivos de entrada
diretorio_entrada = 'D:/Atividades Faculdade/APS/dados/lgndform'

# Criar uma nova pasta de trabalho
wb = Workbook()
# Selecionar a planilha ativa
ws = wb.active

# Adicionando um título para a coluna
ws['A1'] = 'Legenda'

# Inicializando o contador de linha
linha = 2

# Iterando sobre os arquivos na pasta de entrada
for arquivo in os.listdir(diretorio_entrada):
    if linha <= 100:
        if arquivo.endswith('.txt'):
            # Caminho completo do arquivo de entrada
            caminho_arquivo_entrada = os.path.join(diretorio_entrada, arquivo)
            with open(caminho_arquivo_entrada, 'r', encoding='utf-8') as file:
                # Lendo todo o conteúdo do arquivo em uma única linha
                texto = file.read()
                # Adicionando o conteúdo do arquivo como uma linha na planilha
                ws.cell(row=linha, column=1, value=texto.strip())
                # Incrementando o contador de linha para a próxima linha
                linha += 1

# Salvando a planilha Excel
wb.save('PlanilhaTreinoInicial.xlsx')